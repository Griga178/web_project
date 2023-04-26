from openpyxl import Workbook, load_workbook

def list_to_excel(py_list, excel_path):
    wb = Workbook()
    current_sheet = wb.active
    for el in py_list:
        current_sheet.append(el)
    wb.save(excel_path)

def excel_to_list(xl_name, **kwargs):
    wb = load_workbook(xl_name, read_only = True, data_only = True)
    sheet_name = kwargs['sheet_name'] if 'sheet_name' in kwargs else wb.sheetnames[0]
    min_row = 2 if kwargs.get('headers') == False else 1
    active_sheet = wb[sheet_name]
    column_indices = []
    if kwargs.get('headers_names'):
        for column_name in active_sheet[1]:
            if column_name.value in kwargs['headers_names']:
                column_index = active_sheet[1].index(column_name)
                column_indices.append(column_index)
                kwargs['headers_names'].remove(column_name.value)
        else:
            if kwargs["headers_names"]:
                print(f'НЕ НАШЛОСЬ КОЛОНКА(И) {kwargs["headers_names"]}')
    # headers_indexes =
    return_list = []
    for row in active_sheet.iter_rows(min_row = min_row, values_only = True):
        return_row = []
        if column_indices:
            for clm_index in column_indices:
                return_row.append(row[clm_index])
        else:
            return_row = list(row)
        return_list.append(return_row)

    return return_list
