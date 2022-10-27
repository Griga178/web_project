from openpyxl import load_workbook


def get_first_excel_list(file_obj):
    excel_data = load_workbook(file_obj, read_only = True, data_only = True)
    first_excel_list = excel_data.worksheets[0]
    return first_excel_list.iter_rows()
